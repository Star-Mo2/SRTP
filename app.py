"""
============================================================
 Flask Web 应用 v3.0
 三层贝叶斯网络 + localStorage API Key
============================================================
"""
import os, csv, io
from flask import (Flask, render_template, request, jsonify, Response)
from engine import get_engine, rebuild_engine, BNFacilityEngine
from database import (save_evaluation, get_all_evaluations, get_evaluation,
                      delete_evaluation, clear_all, update_interpretation, get_statistics)
from llm_client import (get_system_prompt, save_system_prompt,
                        call_llm, test_llm_connection, get_admin_password,
                        get_knowledge_base, save_knowledge_base)
from rotation import generate_rotation_plan

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "srtp-bn-facility-2026")
_engine = get_engine()

FACILITY_TYPES = ["长椅","儿童滑梯","健身器材","晾衣架","凉亭","护栏",
                  "乒乓球桌","信息公告栏","无障碍坡道",
                  "减速带","路灯","可拆卸路牌","遮阳棚"]


def _safe_pct(val, default=0.0):
    """安全地将值转换为百分比字符串，防止非数字值导致格式化异常"""
    try:
        return f"{float(val):.1%}"
    except (ValueError, TypeError):
        return f"{float(default):.1%}"


def _safe_num(val, default=0.0):
    """安全地将值转换为保留两位小数的字符串"""
    try:
        return f"{float(val):.2f}"
    except (ValueError, TypeError):
        return f"{float(default):.2f}"


def _result_to_row(r):
    return {
        "risk_level":r["risk_level"],"risk_score":r["risk_score"],
        "prob_low":f"{r['prob_low']:.1%}","prob_med":f"{r['prob_med']:.1%}",
        "prob_high":f"{r['prob_high']:.1%}",
        "prob_low_val":r["prob_low"],"prob_med_val":r["prob_med"],"prob_high_val":r["prob_high"],
        "social_weight":r["social_weight"],"weight_tier":r["weight_tier"],
        "priority_score":r["priority_score"],"priority_level":r["priority_level"],
        "exposure":r["exposure"],"usage":r["usage"],
        "maintenance":r["maintenance"],"social_impact":r["social_impact"],
        "mid_probs":r.get("mid_probs",{}),
    }


# ================================================================
#  页面路由
# ================================================================

@app.route("/")
def index():
    """新首页：维护优先级排序"""
    return render_template("index.html", facility_types=FACILITY_TYPES)


@app.route("/diagnose")
def diagnose_page():
    """单设施详细诊断"""
    return render_template("diagnose.html", facility_types=FACILITY_TYPES)

@app.route("/cpt")
def cpt_page():
    return render_template("cpt.html", graph=get_engine().get_network_graph())

@app.route("/history")
def history_page():
    return render_template("history.html")

@app.route("/history/<int:record_id>")
def history_detail(record_id):
    record = get_evaluation(record_id)
    if not record: return "记录不存在", 404
    return render_template("history_detail.html", record=record)

@app.route("/ipa")
def ipa_page():
    return render_template("ipa.html", facility_types=FACILITY_TYPES)

@app.route("/tools")
def tools_page():
    return render_template("tools.html")

@app.route("/calibrate")
def calibrate_page():
    """模型校准页面"""
    return render_template("calibrate.html")

@app.route("/rotation")
def rotation_page():
    """动态轮换建议页面"""
    return render_template("rotation.html", facility_types=FACILITY_TYPES)

@app.route("/admin/login")
def admin_login_page():
    return render_template("admin_login.html")


@app.route("/api/admin/login", methods=["POST"])
def api_admin_login():
    data = request.get_json()
    if data.get("password","") == get_admin_password():
        return jsonify({"success":True})
    return jsonify({"success":False, "error":"密码错误"}), 401


@app.route("/admin")
def admin_page():
    engine = get_engine()
    return render_template("admin.html",
                           system_prompt=get_system_prompt(),
                           knowledge_base=get_knowledge_base(),
                           weights=engine.factor_weights)


# ================================================================
#  评估 API（11 字段）
# ================================================================

@app.route("/api/evaluate", methods=["POST"])
def api_evaluate():
    data = request.get_json()
    engine = get_engine()
    result = engine.evaluate(
        material=data.get("material","金属"),
        install_age=data.get("install_age","3-8年"),
        water_log=data.get("water_log","中"),
        sun_shade=data.get("sun_shade","有遮"),
        use_freq=data.get("use_freq","中"),
        user_group=data.get("user_group","成人"),
        use_intensity=data.get("use_intensity","静坐"),
        inspect_freq=data.get("inspect_freq","每月"),
        repair_time=data.get("repair_time","3-14天"),
        dependency=data.get("dependency","中"),
        outage_impact=data.get("outage_impact","中等"),
        facility_type=data.get("facility_type","长椅"),
        user_groups=data.get("user_groups",["老人"]),
        facility_name=data.get("facility_name",""),
    )
    record_id = save_evaluation(result)
    return jsonify({"success":True, "record_id":record_id, "result":_result_to_row(result), "raw":result})


def _extract_key_factors(result, engine):
    """从中间因子概率中提取关键影响因素简述"""
    factors = []
    mid = result.get("mid_probs", {})
    # 暴露相关
    exp = mid.get("exposure_prob", {})
    if exp.get("高暴露", 0) > exp.get("低暴露", 0):
        factors.append("老化严重")
    # 使用相关
    use = mid.get("usage_prob", {})
    if use.get("高负荷", 0) > use.get("低负荷", 0):
        factors.append("使用过载")
    # 维护相关
    maint = mid.get("maintenance_prob", {})
    if maint.get("维护滞后", 0) > maint.get("维护良好", 0):
        factors.append("维护滞后")
    # 社会影响相关
    soc = mid.get("social_prob", {})
    if soc.get("高影响", 0) > soc.get("低影响", 0):
        factors.append("社会影响大")
    return " + ".join(factors) if factors else "暂无突出风险"


@app.route("/api/ranking/evaluate", methods=["POST"])
def api_ranking_evaluate():
    """批量评估：接收设施列表 → 逐个BN推理 → 按优先级降序排列"""
    try:
        data = request.get_json()
        facilities = data.get("facilities", [])
        if not facilities:
            return jsonify({"success": False, "error": "设施列表为空"}), 400

        engine = get_engine()
        defaults = {
            "facility_type": "长椅", "facility_name": "",
            "material": "金属", "install_age": "3-8年", "water_log": "中", "sun_shade": "有遮",
            "use_freq": "中", "user_group": "成人", "use_intensity": "静坐",
            "inspect_freq": "每月", "repair_time": "3-14天",
            "dependency": "中", "outage_impact": "中等",
            "user_groups": ["老人"],
        }
        # 合法的字段名白名单
        VALID_KEYS = set(defaults.keys())

        results = []
        for i, fac in enumerate(facilities):
            # 合并默认值（只取合法字段，过滤前端可能夹带的 UI 属性）
            params = dict(defaults)
            for k in VALID_KEYS:
                if k in fac and fac[k] is not None and fac[k] != '':
                    params[k] = fac[k]
            if not params["facility_name"]:
                params["facility_name"] = f"{params['facility_type']}-{i+1}"

            r = engine.evaluate(
                material=params["material"], install_age=params["install_age"],
                water_log=params["water_log"], sun_shade=params["sun_shade"],
                use_freq=params["use_freq"], user_group=params["user_group"],
                use_intensity=params["use_intensity"],
                inspect_freq=params["inspect_freq"], repair_time=params["repair_time"],
                dependency=params["dependency"], outage_impact=params["outage_impact"],
                facility_type=params["facility_type"], user_groups=params["user_groups"],
                facility_name=params["facility_name"],
            )
            key_factors = _extract_key_factors(r, engine)
            results.append({
                **{k: r[k] for k in ["facility_name", "facility_type", "risk_level",
                    "risk_score", "social_weight", "priority_score", "priority_level",
                    "prob_low", "prob_med", "prob_high",
                    "exposure", "usage", "maintenance", "social_impact"]},
                "user_groups": r["user_groups"],
                "key_factors": key_factors,
                "_params": {k: r[k] for k in [
                    "material", "install_age", "water_log", "sun_shade",
                    "use_freq", "user_group", "use_intensity",
                    "inspect_freq", "repair_time", "dependency", "outage_impact",
                ]},
            })

        results.sort(key=lambda x: x["priority_score"], reverse=True)
        for i, r in enumerate(results):
            r["rank"] = i + 1

        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": f"服务器错误: {str(e)}", "traceback": traceback.format_exc()}), 500


@app.route("/api/evaluate/batch", methods=["POST"])
def api_evaluate_batch():
    data = request.get_json()
    engine = get_engine()
    defaults = {k: data.get(k, v) for k,v in [
        ("material","金属"),("install_age","3-8年"),("water_log","中"),("sun_shade","有遮"),
        ("use_freq","中"),("user_group","成人"),("use_intensity","静坐"),
        ("inspect_freq","每月"),("repair_time","3-14天"),
        ("dependency","中"),("outage_impact","中等"),
    ]}
    defaults["user_groups"] = data.get("user_groups",["老人"])
    types = data.get("facility_types", FACILITY_TYPES[:5])
    results = []
    for ft in types:
        r = engine.evaluate(**defaults, facility_type=ft, facility_name=ft)
        results.append(_result_to_row(r))
    return jsonify({"success":True, "results":results})


@app.route("/api/what-if", methods=["POST"])
def api_what_if():
    data = request.get_json()
    engine = get_engine()
    current = {k: data[k] for k in ["material","install_age","water_log","sun_shade",
                                      "use_freq","user_group","use_intensity",
                                      "inspect_freq","repair_time",
                                      "dependency","outage_impact"]}
    result = engine.what_if(current, data.get("change_var",""), data.get("new_value",""))
    return jsonify({"success":True, "result":result})


@app.route("/api/scenario/compare", methods=["POST"])
def api_scenario_compare():
    data = request.get_json()
    engine = get_engine()
    result = engine.compare_scenarios(
        material=data.get("material","金属"), install_age=data.get("install_age","3-8年"),
        water_log=data.get("water_log","中"), sun_shade=data.get("sun_shade","有遮"),
        use_freq=data.get("use_freq","中"), user_group=data.get("user_group","成人"),
        use_intensity=data.get("use_intensity","静坐"),
        inspect_freq=data.get("inspect_freq","每月"), repair_time=data.get("repair_time","3-14天"),
        dependency=data.get("dependency","中"), outage_impact=data.get("outage_impact","中等"),
        facility_type=data.get("facility_type","长椅"),
        user_groups=data.get("user_groups",["老人"]),
        facility_name=data.get("facility_name",""),
    )
    return jsonify({"success":True, "scenarios": result})


# ================================================================
#  CPT + 网络图
# ================================================================

@app.route("/api/network/graph")
def api_network_graph():
    return jsonify({"success":True, **get_engine().get_network_graph()})

@app.route("/api/cpt/influence")
def api_cpt_influence():
    return jsonify({"success":True, "influences": get_engine().get_factor_influence()})


# ================================================================
#  IPA
# ================================================================

@app.route("/api/ipa/analyze", methods=["POST"])
def api_ipa_analyze():
    return jsonify({"success":True, **BNFacilityEngine.ipa_analyze(request.get_json().get("facilities",[]))})


# ================================================================
#  文件导入解析
# ================================================================

# 列名模糊匹配表：用户填的列名 → 内部字段名
COLUMN_MAP = {
    # 首页 11 字段
    "材料":"material","材料类型":"material","材质":"material",
    "安装年龄":"install_age","安装年份":"install_age","使用年限":"install_age","年龄":"install_age",
    "积水风险":"water_log","积水":"water_log","排水":"water_log",
    "遮阴":"sun_shade","遮阴情况":"sun_shade","遮阳":"sun_shade","日照":"sun_shade",
    "使用频率":"use_freq","频率":"use_freq",
    "主要群体":"user_group","主要使用群体":"user_group","使用群体":"user_group","群体":"user_group",
    "使用强度":"use_intensity","强度":"use_intensity","使用方式":"use_intensity",
    "巡检频率":"inspect_freq","巡检":"inspect_freq","检查频率":"inspect_freq",
    "维修响应":"repair_time","维修响应时间":"repair_time","响应时间":"repair_time","维修":"repair_time",
    "群体依赖度":"dependency","群体依赖":"dependency","依赖程度":"dependency","依赖":"dependency",
    "无障碍设施":"outage_impact","无障碍":"outage_impact",
    # 停用后影响等级
    "停用后影响":"outage_impact","停用影响":"outage_impact","影响等级":"outage_impact",
    "停用后影响等级":"outage_impact","停用影响等级":"outage_impact",
    # 轮换专用
    "周边同类设施":"same_type_nearby","同类设施距离":"same_type_nearby","替代设施距离":"same_type_nearby",
    "当前健康度":"current_health","健康度":"current_health","当前磨损":"current_health",
    "设施名称":"facility_name","名称":"facility_name",
    "设施类型":"facility_type","类型":"facility_type",
    # 校准中间因子（第2层）——新名对齐申请表，旧名保留兼容
    "退化暴露":"exposure","老化暴露":"exposure","暴露":"exposure","老化":"exposure",
    "使用负荷":"usage","负荷":"usage",
    "治理滞后":"maintenance","维护滞后":"maintenance","维护":"maintenance",
    "正义修正因子":"social_impact","社会影响因素":"social_impact","社会影响":"social_impact",
    "实际故障风险":"risk_level","故障风险":"risk_level","真实风险":"risk_level",
    # IPA 字段
    "重要性":"importance","重要性打分":"importance","重要程度":"importance",
    "满意度":"performance","满意度打分":"performance","满意程度":"performance","表现":"performance",
}

ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}

def parse_upload(file):
    """解析上传的 CSV 或 Excel 文件，返回 [{colname: value}]"""
    import openpyxl
    filename = file.filename or ''
    ext = filename[filename.rfind('.'):].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"不支持的文件格式: {ext}。请上传 CSV 或 .xlsx 文件。")

    rows = []
    if ext == '.csv':
        text = file.read().decode('utf-8-sig')
        lines = [l for l in text.splitlines() if l.strip()]
        if not lines: raise ValueError("文件为空")
        headers = [h.strip() for h in lines[0].split(',')]
        for line in lines[1:]:
            vals = [v.strip() for v in line.split(',')]
            rows.append(dict(zip(headers, vals)))
    else:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v or '').strip() for v in row])))
        wb.close()

    if not rows: raise ValueError("文件中没有数据行")

    # 列名模糊匹配
    import re
    parsed = []
    for row in rows:
        mapped = {}
        for col, val in row.items():
            key = COLUMN_MAP.get(col, None)
            if key is None:
                # 去掉列名中的括号提示（如 "老化暴露(高/中/低)" → "老化暴露"）
                clean = re.sub(r'[（(][^)）]*[)）]', '', col).strip()
                key = COLUMN_MAP.get(clean, None)
            if key is None:
                # 模糊匹配：去掉空格后比较
                clean = col.replace(' ', '').replace('_', '').lower()
                for ck, cv in COLUMN_MAP.items():
                    if ck.replace(' ', '') == clean:
                        key = cv; break
            if key: mapped[key] = val
        if mapped: parsed.append(mapped)

    if not parsed: raise ValueError("未能识别任何列名。请检查文件表头是否与模板一致。")
    return {"rows": parsed, "count": len(parsed),
            "message": f"成功解析 {len(parsed)} 条记录。请检查并修正后提交。"}

@app.route("/api/import/parse", methods=["POST"])
def api_import_parse():
    try:
        file = request.files.get('file')
        if not file: return jsonify({"success":False,"error":"未收到文件"}), 400
        result = parse_upload(file)
        return jsonify({"success":True, **result})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)}), 400


# ================================================================
#  CPT 校准
# ================================================================

@app.route("/api/engine/priors", methods=["GET"])
def api_engine_priors():
    """返回当前模型的先验概率（用于校准前快照）"""
    from engine import BNFacilityEngine
    engine = get_engine()
    return jsonify({
        "success": True,
        "priors": engine.get_priors(),
        "labels": engine.LABELS_MAP,
    })


@app.route("/api/calibrate", methods=["POST"])
def api_calibrate():
    """执行校准并返回前后对比"""
    engine = get_engine()
    priors_before = engine.get_priors()
    result = engine.calibrate_from_data(request.get_json().get("records", []))
    if result.get("success"):
        result["priors_before"] = priors_before
    return jsonify(result)


@app.route("/api/engine/reset", methods=["POST"])
def api_engine_reset():
    """重置模型为默认参数"""
    rebuild_engine()
    return jsonify({"success": True, "message": "模型已重置为默认参数"})


# ================================================================
#  动态轮换 API
# ================================================================

@app.route("/api/rotation/plan", methods=["POST"])
def api_rotation_plan():
    """
    接收同类型设施列表 → 每个跑 BN 拿 exposure/usage →
    轮换算法生成方案。
    """
    try:
        data = request.get_json()
        sites = data.get("sites", [])
        T_min = int(data.get("T_min", 3))

        if not sites or len(sites) < 2:
            return jsonify({"success": False, "error": "至少需要 2 个同类型设施"}), 400

        engine = get_engine()
        enriched = []
        for s in sites:
            # 如果填了 same_type_nearby 且没手动填 social 字段，则自动推断
            nearby = s.get("same_type_nearby", "")
            if nearby:
                infer = BNFacilityEngine.infer_social_from_nearby(nearby)
                if not s.get("outage_impact"):
                    s["outage_impact"] = infer.get("outage_impact", "中等")
                if not s.get("dependency"):
                    s["dependency"] = infer.get("dependency", "中")

            # 如果没填 health，则从 install_age + material 推断
            if not s.get("current_health"):
                s["current_health"] = BNFacilityEngine.infer_health_from_age(
                    s.get("install_age", "3-8年"),
                    s.get("material", "金属"),
                )

            # 跑 BN
            r = engine.evaluate(
                material=s.get("material", "金属"),
                install_age=s.get("install_age", "3-8年"),
                water_log=s.get("water_log", "中"),
                sun_shade=s.get("sun_shade", "有遮"),
                use_freq=s.get("use_freq", "中"),
                user_group=s.get("user_group", "成人"),
                use_intensity=s.get("use_intensity", "静坐"),
                inspect_freq=s.get("inspect_freq", "每月"),
                repair_time=s.get("repair_time", "3-14天"),
                dependency=s.get("dependency", "中"),
                outage_impact=s.get("outage_impact", "中等"),
                facility_type=s.get("facility_type", "长椅"),
                user_groups=s.get("user_groups", ["老人"]),
                facility_name=s.get("facility_name", ""),
            )
            enriched.append({
                "facility_name": s.get("facility_name", ""),
                "facility_type": s.get("facility_type", "长椅"),
                "material": s.get("material", "金属"),
                "install_age": s.get("install_age", "3-8年"),
                "sun_shade": s.get("sun_shade", "有遮"),
                "water_log": s.get("water_log", "中"),
                "use_freq": s.get("use_freq", "中"),
                "user_group": s.get("user_group", "成人"),
                "use_intensity": s.get("use_intensity", "静坐"),
                # v1.1: 传递连续概率分布，而非离散标签
                "exposure": r["exposure"],
                "exposure_probs": r["mid_probs"]["exposure_prob"],
                "usage": r["usage"],
                "usage_probs": r["mid_probs"]["usage_prob"],
                "same_type_nearby": s.get("same_type_nearby", ""),
                "current_health": int(s.get("current_health", 3)),
                "_bn_result": r,
            })

        plan = generate_rotation_plan(enriched, T_min)

        if not plan.get("success"):
            return jsonify(plan), 400

        return jsonify({
            "success": True,
            "plan": plan,
            "sites": enriched,
        })
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/rotation/import", methods=["POST"])
def api_rotation_import():
    """
    轮换专用导入：解析 CSV/XLSX → 检测设施类型 →
    如果含多种类型 → 返回类型列表让用户选 →
    用户选完后回传，只导入选定类型。
    """
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"success": False, "error": "未收到文件"}), 400

        result = parse_upload(file)
        rows = result["rows"]
        if not rows:
            return jsonify({"success": False, "error": "文件中没有有效数据"}), 400

        # 检测设施类型
        type_set = set()
        for r in rows:
            ft = r.get("facility_type", "").strip()
            if ft:
                type_set.add(ft)

        if len(type_set) == 0:
            return jsonify({"success": False, "error": "文件中未检测到「设施类型」列"}), 400

        # 检查无效值（轮换字段合法值）
        VALID_NEARBY = ["近(<50m)", "中(50-200m)", "远(>200m)"]
        HEALTH_RANGE = ["1", "2", "3", "4", "5"]
        invalid_rows = []
        for i, r in enumerate(rows):
            issues = []
            nb = r.get("same_type_nearby", "").strip()
            if nb and nb not in VALID_NEARBY:
                issues.append(f"周边同类设施值「{nb}」无效")
            hl = r.get("current_health", "").strip()
            if hl and hl not in HEALTH_RANGE:
                issues.append(f"当前健康度值「{hl}」无效（应填 1-5）")
            if issues:
                invalid_rows.append({"index": i + 2, "issues": issues, "name": r.get("facility_name", "")})

        return jsonify({
            "success": True,
            "types": sorted(list(type_set)),
            "multi_type": len(type_set) > 1,
            "count": len(rows),
            "rows": rows,
            "invalid_rows": invalid_rows,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@app.route("/api/rotation/import-filtered", methods=["POST"])
def api_rotation_import_filtered():
    """
    用户在类型选择后，回传选定的类型 → 只返回该类型的行。
    同时做无效值标记（⚠️ 前缀）。
    """
    try:
        data = request.get_json()
        rows = data.get("rows", [])
        selected_type = data.get("selected_type", "").strip()

        if not selected_type:
            return jsonify({"success": False, "error": "未选择设施类型"}), 400

        VALID_NEARBY = ["近(<50m)", "中(50-200m)", "远(>200m)"]
        HEALTH_RANGE = ["1", "2", "3", "4", "5"]

        filtered = []
        warnings = []
        for i, r in enumerate(rows):
            ft = r.get("facility_type", "").strip()
            if ft != selected_type:
                continue

            # 无效值标记
            nb = r.get("same_type_nearby", "").strip()
            hl = r.get("current_health", "").strip()
            flags = []
            if nb and nb not in VALID_NEARBY:
                flags.append(f"⚠️ 周边同类设施「{nb}」无效")
            if hl and hl not in HEALTH_RANGE:
                flags.append(f"⚠️ 当前健康度「{hl}」无效（应填 1-5 整数）")
            if flags:
                warnings.append({"index": i + 2, "name": r.get("facility_name", ""), "flags": flags})

            # 推断缺失的 social 字段
            if nb and nb in VALID_NEARBY:
                infer = BNFacilityEngine.infer_social_from_nearby(nb)
                if not r.get("outage_impact", "").strip():
                    r["outage_impact"] = infer.get("outage_impact", "")
                if not r.get("dependency", "").strip():
                    r["dependency"] = infer.get("dependency", "")

            # 推断缺失的 health
            if not hl:
                r["current_health"] = str(BNFacilityEngine.infer_health_from_age(
                    r.get("install_age", "3-8年"),
                    r.get("material", "金属"),
                ))

            filtered.append(r)

        return jsonify({
            "success": True,
            "selected_type": selected_type,
            "count": len(filtered),
            "rows": filtered,
            "warnings": warnings,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ================================================================
#  LLM API（Key 从请求传入，服务端不保存）
# ================================================================

@app.route("/api/llm/interpret", methods=["POST"])
def api_llm_interpret():
    data = request.get_json()
    api_key = data.get("api_key","")
    endpoint = data.get("endpoint","https://api.deepseek.com/v1/chat/completions")
    model = data.get("model","deepseek-chat")

    record_id = data.get("record_id")
    record = get_evaluation(record_id) if record_id else None

    if record:
        info = record
    else:
        info = data.get("result", {})

    user_message = f"""请根据以下老旧小区适老化设施维护优先级评估结果，给出专业的解读和维护建议：

【设施信息】
- 设施名称：{info.get('facility_name','未命名')}
- 设施类型：{info.get('facility_type','')}
- 主要使用群体：{info.get('user_groups',[])}

【观测变量】
- 材料：{info.get('material','')} | 安装年龄：{info.get('install_age','')}
- 积水风险：{info.get('water_log','')} | 遮阴：{info.get('sun_shade','')}
- 使用频率：{info.get('use_freq','')} | 主要群体：{info.get('user_group','')}
- 使用强度：{info.get('use_intensity','')}
- 巡检频率：{info.get('inspect_freq','')} | 维修响应：{info.get('repair_time','')}
- 群体依赖度：{info.get('dependency','')} | 停用后影响等级：{info.get('outage_impact','')}

【中间推断结果】
- 老化暴露：{info.get('exposure','')} | 使用负荷：{info.get('usage','')}
- 维护状态：{info.get('maintenance','')} | 社会影响：{info.get('social_impact','')}

【最终评估】
- 故障风险等级：{info.get('risk_level','')}
- 低/中/高风险概率：{_safe_pct(info.get('prob_low',0))} / {_safe_pct(info.get('prob_med',0))} / {_safe_pct(info.get('prob_high',0))}
- 风险得分：{_safe_num(info.get('risk_score',0))}/3.0 | 社会权重：{_safe_num(info.get('social_weight',0))}
- 优先级得分：{_safe_num(info.get('priority_score',0))} | 等级：{info.get('priority_level','')}

请从以下角度给出建议：
1. 当前风险状况的解读
2. 最关键的影响因素是什么
3. 具体的维护建议（近期应急 + 长期预防）
4. 如果不及时维护可能产生的后果"""

    try:
        interpretation = call_llm(api_key, endpoint, model, get_system_prompt(), user_message)
        if record_id: update_interpretation(record_id, interpretation)
        return jsonify({"success":True, "interpretation":interpretation})
    except Exception as e:
        return jsonify({"success":False, "error":str(e)}), 500


@app.route("/api/llm/test", methods=["POST"])
def api_llm_test():
    data = request.get_json()
    success, msg = test_llm_connection(
        data.get("api_key",""), data.get("endpoint",""), data.get("model",""))
    return jsonify({"success":success, "message":msg})


# ================================================================
#  历史 API
# ================================================================

@app.route("/api/history/list")
def api_history_list():
    page = request.args.get("page",1,type=int)
    data = get_all_evaluations(page=page, risk_filter=request.args.get("risk"),
                               priority_filter=request.args.get("priority"))
    for r in data["records"]:
        r["prob_low_str"]=f"{r['prob_low']:.1%}"
        r["prob_med_str"]=f"{r['prob_med']:.1%}"
        r["prob_high_str"]=f"{r['prob_high']:.1%}"
    return jsonify({"success":True, **data})

@app.route("/api/history/<int:record_id>", methods=["DELETE"])
def api_history_delete(record_id):
    delete_evaluation(record_id)
    return jsonify({"success":True})

@app.route("/api/history/clear", methods=["DELETE"])
def api_history_clear():
    clear_all()
    return jsonify({"success":True})

@app.route("/api/history/stats")
def api_history_stats():
    return jsonify({"success":True, **get_statistics()})

@app.route("/api/export/csv")
def api_export_csv():
    data = get_all_evaluations(per_page=99999)
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["时间","设施名称","类型","使用群体",
                "材料","安装年龄","积水","遮阴",
                "使用频率","主要群体","使用强度",
                "巡检频率","维修响应","群体依赖","停用后影响",
                "暴露","负荷","维护","社会影响",
                "风险等级","低风险%","中风险%","高风险%",
                "风险得分","社会权重","优先级得分","优先级等级"])
    for r in data["records"]:
        w.writerow([r["created_at"],r["facility_name"],r["facility_type"],
                    ", ".join(r["user_groups"]),
                    r.get("material",""),r.get("install_age",""),r.get("water_log",""),r.get("sun_shade",""),
                    r.get("use_freq",""),r.get("user_group",""),r.get("use_intensity",""),
                    r.get("inspect_freq",""),r.get("repair_time",""),r.get("dependency",""),r.get("outage_impact",""),
                    r.get("exposure",""),r.get("usage",""),r.get("maintenance",""),r.get("social_impact",""),
                    r["risk_level"],r["prob_low"],r["prob_med"],r["prob_high"],
                    r["risk_score"],r["social_weight"],r["priority_score"],r["priority_level"]])
    output.seek(0)
    return Response(output.getvalue().encode("utf-8-sig"), mimetype="text/csv",
                    headers={"Content-Disposition":"attachment; filename=evaluations_export.csv"})


# ================================================================
#  管理员 API（无需密码）
# ================================================================

@app.route("/api/admin/prompt", methods=["POST"])
def api_admin_prompt():
    save_system_prompt(request.get_json().get("prompt",""))
    return jsonify({"success":True})

@app.route("/api/admin/knowledge-base", methods=["GET"])
def api_admin_get_knowledge_base():
    return jsonify({"success":True, "knowledge_base": get_knowledge_base()})

@app.route("/api/admin/knowledge-base", methods=["POST"])
def api_admin_save_knowledge_base():
    save_knowledge_base(request.get_json().get("knowledge_base",""))
    return jsonify({"success":True})

@app.route("/api/admin/weights", methods=["POST"])
def api_admin_weights():
    rebuild_engine(request.get_json())
    return jsonify({"success":True, "weights": get_engine().factor_weights})

@app.route("/api/admin/weights", methods=["GET"])
def api_admin_get_weights():
    return jsonify({"success":True, "weights": get_engine().factor_weights})


# ================================================================
#  启动
# ================================================================

if __name__ == "__main__":
    print("\n" + "=" * 50)
    print("  老旧小区适老化设施维护优先级评估系统 v3.0")
    print("  三层贝叶斯网络 | API Key 浏览器本地存储")
    print("  http://127.0.0.1:5000")
    print("=" * 50 + "\n")
    app.run(host="127.0.0.1", port=5000, debug=os.environ.get("FLASK_DEBUG", "0") == "1")
